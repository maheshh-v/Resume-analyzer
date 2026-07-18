"""Parsing a candidate sheet (CSV or XLSX) into rows the import endpoint can act on.

Header matching is forgiving — "GitHub username", "github_login", and "github" all map to the
same field — because these files come straight out of whatever ATS export or spreadsheet the
customer already has. Bad rows are reported by their spreadsheet row number and never block
the good ones.
"""

import csv
import io
from dataclasses import dataclass, field

MAX_SHEET_ROWS = 200

_HEADER_ALIASES: dict[str, set[str]] = {
    "name": {"name", "full name", "fullname", "candidate", "candidate name"},
    "email": {"email", "e-mail", "mail", "email address"},
    "github_login": {"github", "github login", "github_login", "github username", "github handle", "github user"},
    "linkedin_url": {"linkedin", "linkedin url", "linkedin_url", "linkedin profile"},
    "resume_url": {"resume url", "resume_url", "resume link", "resume", "cv url", "cv_url", "cv link", "cv"},
}


@dataclass
class SheetRow:
    row_number: int  # 1-based as the user sees it in their spreadsheet (header is row 1)
    name: str
    email: str | None = None
    github_login: str | None = None
    linkedin_url: str | None = None
    resume_url: str | None = None


@dataclass
class SheetParseResult:
    rows: list[SheetRow] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _canonical_field(header: str) -> str | None:
    normalized = header.strip().lower()
    for field_name, aliases in _HEADER_ALIASES.items():
        if normalized in aliases:
            return field_name
    return None


def _cells_from_csv(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    return [[cell.strip() for cell in row] for row in csv.reader(io.StringIO(text))]


def _cells_from_xlsx(content: bytes) -> list[list[str]]:
    from openpyxl import load_workbook  # local import: only paid on xlsx uploads

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        return [["" if v is None else str(v).strip() for v in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def parse_candidate_sheet(content: bytes, filename: str) -> SheetParseResult:
    result = SheetParseResult()
    lowered = filename.lower()
    try:
        if lowered.endswith(".csv"):
            cells = _cells_from_csv(content)
        elif lowered.endswith(".xlsx"):
            cells = _cells_from_xlsx(content)
        else:
            result.errors.append("Unsupported file type — upload a .csv or .xlsx file")
            return result
    except Exception:
        result.errors.append("Couldn't read the file — is it a valid CSV/XLSX?")
        return result

    non_empty = [(i, row) for i, row in enumerate(cells) if any(cell for cell in row)]
    if not non_empty:
        result.errors.append("The sheet is empty")
        return result

    header_index, header_row = non_empty[0]
    column_map = {col: _canonical_field(header) for col, header in enumerate(header_row)}
    if "name" not in column_map.values():
        result.errors.append('No "name" column found — the sheet needs at least a name header')
        return result

    data_rows = non_empty[1:]
    if len(data_rows) > MAX_SHEET_ROWS:
        result.errors.append(f"Sheet has more than {MAX_SHEET_ROWS} data rows — split it into smaller files")
        data_rows = data_rows[:MAX_SHEET_ROWS]

    for index, row in data_rows:
        row_number = index + 1  # spreadsheet numbering, matching what the user sees
        values: dict[str, str] = {}
        for col, field_name in column_map.items():
            if field_name and col < len(row) and row[col]:
                values[field_name] = row[col]

        name = values.get("name", "").strip()
        if not name:
            result.errors.append(f"Row {row_number}: missing name")
            continue
        email = values.get("email")
        if email and "@" not in email:
            result.errors.append(f"Row {row_number}: '{email}' doesn't look like an email")
            continue
        resume_url = values.get("resume_url")
        if resume_url and not resume_url.lower().startswith(("http://", "https://")):
            result.errors.append(f"Row {row_number}: resume URL must start with http(s)://")
            continue

        result.rows.append(
            SheetRow(
                row_number=row_number,
                name=name,
                email=email,
                github_login=values.get("github_login"),
                linkedin_url=values.get("linkedin_url"),
                resume_url=resume_url,
            )
        )
    return result
